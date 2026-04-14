"""
school_coverage_assessment.py
------------------------------
External validity assessment for the LAC school accessibility platform.

For each country this script computes:
  1. Coverage  — our processed school count as % of the national universe
                 (public-only universe and total public+private universe)
  2. Georeferencing rate — % of schools in our file that have valid coordinates

It also patches every processed CSV with two new columns:
  - year       : reference year of the school data
  - id_national: country-level school identifier (mirrors id_centro)

Outputs
-------
  results/school_coverage_assessment.csv   — machine-readable summary
  results/school_coverage_assessment.xlsx  — formatted Excel workbook

Usage
-----
  uv run python school_coverage_assessment.py
  # or plain:
  python school_coverage_assessment.py
"""

import os
import re
import warnings
import pandas as pd
from pathlib import Path

warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
# 1. COUNTRY METADATA
#    Each entry:  ISO → (year, public_universe, total_universe, sector_scope, notes)
#
#    public_universe  = schools matching R script's public/official filter
#                       (counted directly from each country's ministry raw frame file,
#                        applying the same filters as the R/Quarto pipeline)
#    total_universe   = all schools in raw frame at same education levels (public + private)
#    sector_scope     = what the R script covers: "public", "public+aided", "all", etc.
#
#    ALL figures derived programmatically from the raw ministry frame files in
#    data/schools/AR/{ISO}/raw/ — not from hardcoded estimates or UNESCO UIS.
#    A _ministry_counts.json file in each country's raw/ folder documents the derivation.
#    Updated: 2026-03-19
#
#    universe_data_source:
#      "ministry_raw"  — counted directly from the raw ministry frame file in this repo
#      "incomplete"    — raw file exists but is known to be partial (OSM scrape, tiny shapefile)
#      "no_raw_file"   — no raw file available; universe from processed CSV or prior estimate
# ──────────────────────────────────────────────────────────────────────────────
COUNTRY_META = {
    # ISO: (data_year, public_universe, total_universe, sector_scope, universe_data_source, source_notes)
    # All figures below are derived directly from each country's ministry raw frame file
    # (same file the R/Quarto pipeline reads), applying the same filters as the R script.
    # 'public_universe'  = schools matching R script's public/official filter
    # 'total_universe'   = all schools in the raw frame (public + private), same education levels

    "ARG": (2024, 33_945, 45_000, "public", "ministry_raw",
            "SOURCE: raw/6831-Listado establecimientos.csv. "
            "Filter: comun=='X' (regular education) & cueanexo ends in '00' (main annexes). "
            "Total=45,000 main regular establishments; sector=1(estatal)=33,945, "
            "sector=2(privado)=10,670, sector=3(cooperativa)=385."),

    "BHS": (2024, 77, 77, "public", "incomplete",
            "SOURCE: raw/bhs_schools_shp/bhs_schools.shp. "
            "77 school POIs — government directory only (no private schools in file). "
            "True universe likely larger; file appears to be public schools only."),

    "BLZ": (2024, 303, 303, "public+aided", "ministry_raw",
            "SOURCE: raw/geo_schools Belize.xlsx. "
            "303 schools, all classified as Government (74) or Government Aided (224) "
            "or Specially Assisted (5). No private schools in file — file is public/aided only. "
            "Script applies no sector filter."),

    "BOL": (2023, 15_329, 16_159, "public", "ministry_raw",
            "SOURCE: raw/MinEdu_InstitucionesEducativas_2023.xlsx (skip=7). "
            "Filter: sub_sistema=='Regular'. Total=16,159 unique codigo_R.U.E. (school registry codes). "
            "Each RUE is one school (multiple RUEs can share the same edificio/building). "
            "dependencia: Fiscal=14,198, Convenio=1,130, Comunitaria=1, Privada=830. "
            "Public (Fiscal+Convenio+Comunitaria)=15,329; Private (Privada)=830 RUE codes."),

    "BRA": (2023, 137_914, 180_230, "public", "ministry_raw",
            "SOURCE: raw/microdados_censo_escolar_2023/dados/microdados_ed_basica_2023.csv. "
            "Filter: tp_situacao_funcionamento==1 (active). Total active=180,230. "
            "tp_dependencia: 1(Federal)=725, 2(Estadual)=33,546, 3(Municipal)=130,709, "
            "4(Privada)=52,645. Public (dep 1-3)=137,914; Private=42,316 active schools. "
            "17,876 public schools lack coordinates (13.3%% — largest gap in LAC)."),

    "BRB": (2024, 106, 106, "unknown", "ministry_raw",
            "SOURCE: raw/Barbados Geolocalización Escuelas 2024.xlsx. "
            "106 records — no sector column in file. Likely government schools only. "
            "CRITICAL: processed file has NO lat/lon column — zero coordinates. Needs fixing."),

    "CHL": (2023, 7_895, 16_659, "public+subvencionado", "ministry_raw",
            "SOURCE: raw/20230912_Directorio_Oficial_EE_2023_20230430_WEB.csv. "
            "Total=16,659 unique RBDs in official directory. "
            "cod_depe: 1(Municipal)=1,149, 2(Part.Subvencionado)=4,846, 3(Part.Pagado)=7,631, "
            "4(Corp.Delegada)=2,114, 5(JUNAEB)=70, 6(SLE)=849. "
            "Script joins directorio to enrollment (filter cod_depe!=4), yielding 7,895. "
            "Truly public (dep 1+6)=1,998; Subvencionado (dep 2, state-funded private)=4,846."),

    "COL": (2023, 42_846, 50_866, "public", "ministry_raw",
            "SOURCE: raw/DANE_2023/Carátula única de la sede educativa.csv. "
            "Filter: estado_id!=0 & novedad_id==9 (active, no outstanding issues). "
            "Total=50,866 active sedes. sector_id=1(Oficial/public)=42,846; "
            "sector_id=2(Privado)=8,020."),

    "CRI": (2024, 4_504, 5_268, "public", "ministry_raw",
            "SOURCE: raw/NominaCentrosEducativos2024.xlsx (sheets: Preescolar, I y II Ciclos, Colegios). "
            "Filter: circuito not null (enrolled schools). Total=5,268 across sheets. "
            "dependencia: PUB=4,504, PRI=724, SUB=40. Script filters to PUB only."),

    "DOM": (2024, 7_893, 10_615, "public", "ministry_raw",
            "SOURCE: raw/RTz-8sq-centros-educativos-2023-2024.csv. "
            "Filter: año==20232024 (most recent year). Total=10,615 centros. "
            "sector: PÚBLICO=7,760, PRIVADO=2,722, SEMIOFICIAL=133. "
            "Public+semiofficial=7,893. R script used 2022-2023 data (7,735 all public)."),

    "ECU": (2024, 12_767, 15_706, "public", "ministry_raw",
            "SOURCE: raw/2_MINEDUC_RegistrosAdministrativos_2024-2025Inicio.csv. "
            "Filter: Tipo Educación=='Ordinaria'. Total Ordinaria=15,706 schools (AMIE unique). "
            "Sostenimiento: Fiscal=12,167, Particular=2,939, Fiscomisional=496, Municipal=104. "
            "Public (non-Particular)=12,767. Script filters sostenimie!='Particular'."),

    "GTM": (2024, 35_497, 35_501, "public", "ministry_raw",
            "SOURCE: raw/sire_2024_filtrado/sire_2024_filtrado.shp (DBF). "
            "Filter: situacion=='ABIERTA'. Total active=35,501 unique school codes. "
            "sector: OFICIAL=34,915, COOPERATIVA=1,242, MUNICIPAL=319, PRIVADO=4. "
            "NOTE: file is already pre-filtered (name='filtrado'); full SIRE with private is larger. "
            "Script filters sector!='PRIVADO'; non-private active=35,497."),

    "GUY": (2024, 874, 874, "public", "ministry_raw",
            "SOURCE: raw/School Data-Mapping.xlsx. "
            "Filter: type!='SEND' (exclude special needs). Total non-SEND=874 schools. "
            "type values: P(Primary)=347, N(Nursery)=328, S(Secondary)=115, PT(Post-primary)=84. "
            "No private schools in file — appears to be government schools only."),

    "HND": (2023, 17_428, 17_428, "public", "ministry_raw",
            "SOURCE: raw/SIPLIE_nivel nacional.xlsx (sheet='Detalle', skip=7). "
            "Filter: nivel!='Básica - Adultos'. Total=17,438 rows → 17,428 unique school codes. "
            "SIPLIE is the government school information system — covers public schools only. "
            "Private schools are not registered in SIPLIE."),

    "HTI": (2022, None, None, "unknown", "incomplete",
            "SOURCE: raw/hti_schools_shp/ (OSM planet points). "
            "OSM file has 5,284 school points — but this is INCOMPLETE. "
            "Raw ministry file haiti-2022-cartographie-des-institutions-scolaires.xls "
            "could not be read (requires xlrd for .xls format). "
            "Known figure: Haiti has ~15,000-18,000 schools (MENFP), ~80-90%% private. "
            "CRITICAL: 0%% platform coverage. Highest-priority country for processing."),

    "JAM": (2024, 955, 955, "unknown", "no_raw_file",
            "SOURCE: Processed file only — JAM_script.qmd is an empty placeholder. "
            "No raw ministry frame available in repository. "
            "955 schools in processed file (estimated public only). "
            "True universe: ~1,050-1,200 total (including ~100+ private). "
            "Needs raw data and proper R script to establish reliable universe."),

    "MEX": (2024, 209_629, 243_842, "public", "ministry_raw",
            "SOURCE: raw/siged_total.csv. "
            "Filter: nivel in [PREESCOLAR, PRIMARIA, SECUNDARIA, MEDIA SUPERIOR]. "
            "Total K-12 levels=243,842. control: PÚBLICO=209,629, PRIVADO=34,213. "
            "Script R: filter(nivel in target & control=='PÚBLICO'). "
            "INICIAL(4,992) + LICENCIATURA(6,478) + POSGRADO(4,387) excluded from count."),

    "PAN": (2024, 3_092, 3_660, "public", "ministry_raw",
            "SOURCE: raw/Marco muestral 19 DE JUNIO 2024.xlsx (skip=3). "
            "Filter: marco_muestral=='SUBSISTEMA REGULAR'. Total regular=3,660. "
            "dependencia: OFICIAL=3,135, PARTICULAR=525. "
            "R script adds: oferta_2023!='NO ESPECIFICADO' & estatus!='IPHE' → 3,092 OFICIAL. "
            "Full PAN_all_schools.geojson (pilot) has 3,617 incl. private with centroid fallback."),

    "PER": (2024, 65_266, 88_019, "public", "ministry_raw",
            "SOURCE: raw/Padron.csv (sep=';', encoding=latin-1). "
            "Filter: niv_mod in ['A2','A3','B0','F0'] (preschool/primary/secondary levels). "
            "Total target levels=88,019 unique cod_mod (school×annex codes). "
            "d_gestion: Pública directa=64,207, Privada=22,764, Pública gestión privada=1,059. "
            "Public (non-Privada)=65,266. Padron year 2024 per file."),

    "PRY": (2023, 8_241, 8_847, "public", "ministry_raw",
            "SOURCE: raw/establecimientos_2023.csv (geographic registry, primary universe). "
            "Total=8,847 unique school codes (all registered schools). "
            "Sector from raw/matriculaciones_*.csv join: Oficial=6,648, Privado=606, "
            "Privado Subvencionado=564, No enrollment match=1,029. "
            "Public (non-Privado, incl. NA/no enrollment match)=8,241. "
            "Processed file (8,162) = schools with both geographic registry AND enrollment data."),

    "SLV": (2024, 5_143, 6_026, "public", "ministry_raw",
            "SOURCE: raw/SLV_coord_EDU.csv (for sector breakdown) + CE_2024 El Salvador.xlsx. "
            "SLV_coord has 6,026 schools: Público=5,143, Privado=883. "
            "CE_2024 has 5,160 rows (active enrollment file — sector filter commented out in script). "
            "Script currently includes all in CE_2024 (no sector filter applied). "
            "WARNING: ~883 private schools may be included in processed output."),

    "SUR": (2024, 953, 958, "public+private", "ministry_raw",
            "SOURCE: raw/Suriname School List_03202024.xlsx. "
            "Filter: education_level does not contain 'special' (removes 25 special ed schools). "
            "Total non-special=958. School_type: O.S.(public)=335, RKBO/EBGS/religious=618, "
            "Particulier(private)=5. Script includes all non-special including Particulier."),

    "URY": (2024, 2_597, 2_597, "public", "ministry_raw",
            "SOURCE: raw/CEIP/CEIP.shp + raw/CES/CES.shp + raw/CETP/CETP.shp. "
            "CEIP (primary): 2,183 total → 2,104 non-special-education. "
            "CES (secondary): 319 total → 307 with valid RUEE code. "
            "CETP (technical): 186 total. Combined=2,597. "
            "All are ANEP (national public education system) — no private schools in dataset."),
}

# ──────────────────────────────────────────────────────────────────────────────
# 2. HELPERS
# ──────────────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent / "data" / "schools" / "AR"

def read_processed(iso: str) -> pd.DataFrame | None:
    """Read processed CSV with automatic encoding detection."""
    path = BASE / iso / "processed" / f"{iso}_total.csv"
    if not path.exists():
        return None
    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            return pd.read_csv(path, encoding=enc, low_memory=False)
        except UnicodeDecodeError:
            continue
    return None


def georef_stats(df: pd.DataFrame) -> tuple[int, int, float]:
    """Return (n_total, n_georef, pct_georef)."""
    if df is None:
        return (0, 0, 0.0)
    lat_col = next((c for c in df.columns if "lat" in c.lower()), None)
    if not lat_col:
        return (len(df), 0, 0.0)
    georef = df[lat_col].notna() & (df[lat_col].astype(str).str.strip() != "")
    n_georef = int(georef.sum())
    return (len(df), n_georef, round(n_georef / len(df) * 100, 1) if len(df) else 0.0)


def patch_processed_csv(iso: str, year: int, df: pd.DataFrame) -> None:
    """
    Add 'year' and 'id_national' columns to the processed CSV if not present.
    Writes back in place (latin-1 safe).
    """
    path = BASE / iso / "processed" / f"{iso}_total.csv"
    if not path.exists():
        return

    changed = False

    if "year" not in df.columns:
        df.insert(1, "year", year)
        changed = True

    if "id_national" not in df.columns and "id_centro" in df.columns:
        idx = df.columns.get_loc("id_centro")
        df.insert(idx + 1, "id_national", df["id_centro"])
        changed = True

    if changed:
        df.to_csv(path, index=False, encoding="utf-8")
        print(f"  [{iso}] patched: added {'year, ' if 'year' in df.columns else ''}id_national")


# ──────────────────────────────────────────────────────────────────────────────
# 3. MAIN ASSESSMENT
# ──────────────────────────────────────────────────────────────────────────────
def main():
    results = []

    for iso, (yr, pub_universe, total_universe, sector_scope, data_src, notes) in COUNTRY_META.items():
        df = read_processed(iso)
        n_total, n_georef, pct_georef = georef_stats(df)
        n_missing = n_total - n_georef

        # Coverage calculations
        cov_pub   = round(n_total / pub_universe   * 100, 1) if pub_universe   and n_total else None
        cov_total = round(n_total / total_universe * 100, 1) if total_universe and n_total else None

        # Patch processed CSV
        if df is not None and yr is not None:
            patch_processed_csv(iso, yr, df)

        results.append({
            "country_iso":             iso,
            "data_year":               yr,
            "sector_scope":            sector_scope,
            "universe_data_source":    data_src,
            # ── Our data ──
            "n_schools_in_file":       n_total if n_total else None,
            "n_georef":                n_georef if n_total else None,
            "n_missing_coords":        n_missing if n_total else None,
            "pct_georef":              pct_georef if n_total else None,
            # ── Universe ──
            "public_universe":         pub_universe,
            "total_universe_est":      total_universe,
            "pct_coverage_vs_public":  cov_pub,
            "pct_coverage_vs_total":   cov_total,
            # ── Flags ──
            "file_exists":             df is not None,
            "has_private":             sector_scope in ("public+private", "public+private?"),
            "private_gap_flag":        (sector_scope == "public") and bool(total_universe and pub_universe and total_universe > pub_universe * 1.05),
            # ── Notes ──
            "assessment_notes":        notes,
        })

    df_out = pd.DataFrame(results)

    # ── Save outputs ──────────────────────────────────────────────────────────
    out_dir = Path(__file__).parent / "results"
    out_dir.mkdir(exist_ok=True)

    csv_path = out_dir / "school_coverage_assessment.csv"
    df_out.to_csv(csv_path, index=False, encoding="utf-8")
    print(f"\nSaved: {csv_path}")

    # Excel with formatting
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coverage Assessment"

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Style header row
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)

        # Colour-code georef% column
        pct_col  = df_out.columns.get_loc("pct_georef") + 1
        cov_col  = df_out.columns.get_loc("pct_coverage_vs_total") + 1
        flag_col = df_out.columns.get_loc("private_gap_flag") + 1
        file_col = df_out.columns.get_loc("file_exists") + 1
        src_col  = df_out.columns.get_loc("universe_data_source") + 1

        red_fill    = PatternFill("solid", fgColor="FFB3B3")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        green_fill  = PatternFill("solid", fgColor="C6EFCE")
        orange_fill = PatternFill("solid", fgColor="FCE4D6")
        blue_fill   = PatternFill("solid", fgColor="D9E8FB")   # ministry_raw
        grey_fill   = PatternFill("solid", fgColor="E2E2E2")   # incomplete / no_raw_file

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            pct    = row[pct_col  - 1].value
            cov    = row[cov_col  - 1].value
            flag   = row[flag_col - 1].value
            exists = row[file_col - 1].value
            src    = row[src_col  - 1].value

            # universe_data_source colouring
            if src == "ministry_raw":
                row[src_col - 1].fill = blue_fill
            elif src in ("incomplete", "no_raw_file"):
                row[src_col - 1].fill = grey_fill

            # Georef colouring
            if pct is not None:
                if pct < 90:
                    row[pct_col - 1].fill = red_fill
                elif pct < 97:
                    row[pct_col - 1].fill = yellow_fill
                else:
                    row[pct_col - 1].fill = green_fill

            # Coverage vs total colouring
            if cov is not None:
                if cov < 60:
                    row[cov_col - 1].fill = red_fill
                elif cov < 80:
                    row[cov_col - 1].fill = yellow_fill
                else:
                    row[cov_col - 1].fill = green_fill

            # Private gap flag
            if flag:
                row[flag_col - 1].fill = orange_fill

            # No file = whole row light red
            if not exists:
                for cell in row:
                    if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = PatternFill("solid", fgColor="FFE0E0")

        # Column widths
        col_widths = {
            "country_iso": 10, "data_year": 9, "sector_scope": 20,
            "universe_data_source": 16,
            "n_schools_in_file": 14, "n_georef": 10, "n_missing_coords": 14,
            "pct_georef": 10, "public_universe": 14, "total_universe_est": 16,
            "pct_coverage_vs_public": 18, "pct_coverage_vs_total": 18,
            "file_exists": 10, "has_private": 10, "private_gap_flag": 14,
            "assessment_notes": 60,
        }
        for i, col_name in enumerate(df_out.columns, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = col_widths.get(col_name, 15)

        # Freeze header
        ws.freeze_panes = "A2"

        # ── Second sheet: quick summary by file status ────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "LAC School Data — External Validity Summary"
        ws2["A1"].font = Font(bold=True, size=13)

        summary_stats = [
            ("Countries with processed file", int(df_out["file_exists"].sum())),
            ("Countries WITHOUT processed file", int((~df_out["file_exists"]).sum())),
            ("Countries with public-only data", int((df_out["sector_scope"] == "public").sum())),
            ("Countries with possible private inclusion", int(df_out["sector_scope"].str.contains("private").sum())),
            ("Countries with private gap flag (>5%% private not covered)", int(df_out["private_gap_flag"].sum())),
            ("Total schools in platform (all countries)", int(df_out["n_schools_in_file"].sum(skipna=True))),
            ("Total public universe (sum of known)", int(df_out["public_universe"].sum(skipna=True))),
            ("Estimated total universe (public+private)", int(df_out["total_universe_est"].sum(skipna=True))),
            ("Countries with georef% < 97%%", int((df_out["pct_georef"] < 97).sum())),
            ("Countries with georef% < 90%%", int((df_out["pct_georef"] < 90).sum())),
            ("Total missing coordinates (all countries)", int(df_out["n_missing_coords"].sum(skipna=True))),
        ]

        ws2["A3"] = "Metric"; ws2["B3"] = "Value"
        ws2["A3"].font = Font(bold=True); ws2["B3"].font = Font(bold=True)
        for i, (label, val) in enumerate(summary_stats, 4):
            ws2[f"A{i}"] = label
            ws2[f"B{i}"] = val
        ws2.column_dimensions["A"].width = 55
        ws2.column_dimensions["B"].width = 12

        xlsx_path = out_dir / "school_coverage_assessment.xlsx"
        wb.save(xlsx_path)
        print(f"Saved: {xlsx_path}")

    except ImportError:
        print("openpyxl not available — Excel output skipped.")

    # ── Print quick console summary ───────────────────────────────────────────
    print("\n" + "=" * 80)
    print(f"{'ISO':4s}  {'Year':4s}  {'In file':>8}  {'Georef%':>7}  {'Cov/Public':>10}  {'Cov/Total':>9}  Scope")
    print("-" * 80)
    for r in results:
        iso   = r["country_iso"]
        yr    = str(r["data_year"]) if r["data_year"] else "?"
        n     = f"{r['n_schools_in_file']:,}" if r["n_schools_in_file"] else "NO FILE"
        geo   = f"{r['pct_georef']}%" if r["pct_georef"] else "N/A"
        cp    = f"{r['pct_coverage_vs_public']}%" if r["pct_coverage_vs_public"] else "?"
        ct    = f"{r['pct_coverage_vs_total']}%" if r["pct_coverage_vs_total"] else "?"
        scope = r["sector_scope"]
        flag  = " ⚠ PRIVATE GAP" if r["private_gap_flag"] else ""
        print(f"{iso:4s}  {yr:4s}  {n:>8}  {geo:>7}  {cp:>10}  {ct:>9}  {scope}{flag}")
    print("=" * 80)
    print(f"\nTotal in platform: {df_out['n_schools_in_file'].sum(skipna=True):,.0f} schools")
    print(f"Est. total LAC universe: {df_out['total_universe_est'].sum(skipna=True):,.0f} schools")


if __name__ == "__main__":
    main()
